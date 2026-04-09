from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
XLSX_PATH = Path(
    r"D:\06. 평생교육기관\[일반시설]+대전광역시교육청\평생교육시설 목록화 자료(대전광역시교육청).xlsx"
)
OUT = ROOT / "script" / "daejeon_crawler" / "output" / "parsed" / "institutions_daejeon_office_public_notice.json"

TYPE_KEYWORDS = {
    "시민사회단체": "civic_lifelong",
    "언론기관": "media_lifelong",
    "원격": "remote_lifelong",
    "지식인력개발": "employment_training",
    "지식·인력개발": "employment_training",
    "사업장": "workplace_lifelong",
    "학교형태": "school_disability_lifelong",
    "장애인평생교육시설": "school_disability_lifelong",
}


def normalize_text(value: str | None) -> str:
    return re.sub(r"[\s()\-_.]", "", (value or "").lower())


def normalize_spacing(value: str | None) -> str:
    return re.sub(r"\s+", " ", (value or "").strip())


def normalize_address(value: str | None) -> str:
    text = normalize_spacing(value)
    text = text.replace("대전광역시", "대전", 1)
    text = text.replace("대전시", "대전", 1)
    text = re.sub(r"^대전\s+", "대전 ", text)
    text = re.sub(r"\s*,\s*", ", ", text)
    return text.strip()


def infer_type(label: str | None) -> str:
    cleaned = normalize_spacing(label).replace(" ", "")
    for keyword, type_id in TYPE_KEYWORDS.items():
        if keyword.replace(" ", "") in cleaned:
            return type_id
    return "other_private_lifelong"


def infer_sigungu(address: str | None) -> str | None:
    if not address:
        return None
    parts = normalize_spacing(address).split()
    if len(parts) >= 2 and parts[0].startswith("대전"):
        return parts[1]
    return None


def load_existing_items() -> list[dict]:
    if not OUT.exists():
        return []
    return json.loads(OUT.read_text(encoding="utf-8"))


def build_items() -> tuple[list[dict], dict[str, int]]:
    workbook = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    existing_items = load_existing_items()
    existing_by_name = {
        normalize_text(item.get("canonical_name") or item.get("display_name") or ""): item
        for item in existing_items
    }

    items: list[dict] = []
    seen_names: set[str] = set()
    reused_ids = 0
    new_ids = 0

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        name, inst_type, _zipcode, road_address, phone = row[:5]
        canonical_name = normalize_spacing(str(name or ""))
        if not canonical_name:
            continue

        normalized_name = normalize_text(canonical_name)
        if normalized_name in seen_names:
            continue
        seen_names.add(normalized_name)

        existing = existing_by_name.get(normalized_name)
        if existing:
            institution_id = existing["institution_id"]
            reused_ids += 1
        else:
            institution_id = f"dj-dje-{len(items) + 1:04d}"
            new_ids += 1

        address = normalize_address(str(road_address or ""))
        phone_text = normalize_spacing(str(phone or ""))

        items.append(
            {
                "institution_id": institution_id,
                "canonical_name": canonical_name,
                "display_name": canonical_name,
                "institution_type": infer_type(str(inst_type or "")),
                "operator_name": None,
                "homepage_url": None,
                "phone": phone_text or None,
                "road_address": address or None,
                "sido": "대전",
                "sigungu": infer_sigungu(address),
                "eupmyeondong": None,
                "operation_status": "운영중",
                "confidence_score": 0.92 if phone_text and address else 0.82,
                "source_count": 1,
                "source_ids": ["daejeon_office_public_notice_20260409"],
                "last_crawled_at": "2026-04-09",
                "last_verified_at": "2026-04-09",
            }
        )

    stats = {
        "excel_rows": len(items),
        "previous_rows": len(existing_items),
        "removed_rows": len(existing_items) - reused_ids,
        "reused_ids": reused_ids,
        "new_ids": new_ids,
    }
    return items, stats


def main() -> None:
    items, stats = build_items()
    OUT.write_text(json.dumps(items, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({"output": str(OUT), "count": len(items), **stats}, ensure_ascii=False))


if __name__ == "__main__":
    main()
